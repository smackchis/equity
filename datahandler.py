from openpyxl import *
import pandas as pd
import sqlalchemy as db

'''
This is a common module for querying, inserting, updating and deleting data in databases
The source of data can be a sql or excel databases
Connections and other configurations of databases are stored in an excel file named config.xlsx 
'''

# READ DATABASE CONFIGS FROM CONFIG EXCEL FILE
configs_wb = load_workbook(filename="C:\\wm\\equity\\CONFIG\\configs.xlsx",data_only=True)
configs_ws = configs_wb['IND_EQ_DB_Configs']
db_host,db_port,db_name,db_user,db_user_password,db_driver = '','','','','',''
row = 1
while configs_ws.cell(row=row, column= 1).value != None:
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_host':
            db_host = str(configs_ws.cell(row=row, column=2).value).strip()
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_port':
            db_port = str(configs_ws.cell(row=row, column=2).value).strip()
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_user':
            db_user = str(configs_ws.cell(row=row, column=2).value).strip()
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_user_password':
            db_user_password = str(configs_ws.cell(row=row, column=2).value).strip()
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_name':
            db_name = str(configs_ws.cell(row=row, column=2).value).strip()
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_driver':
            db_driver = str(configs_ws.cell(row=row, column=2).value).strip()
        row = row + 1

# READ DATA FROM DATABASE AND RETURN A PANDA DATAFRAME OF QUERIED RESULT
def read_db_data(query):
    engine_url = db_driver + '://' + db_user + ':' + db_user_password + '@' + db_host + ':' + db_port + '/' + db_name
    engine = db.create_engine(engine_url)
    data = pd.read_sql_query(query, engine)
    return data

# EXECUTE ANY QUERY IN DB, USUALLY USED FOR INSERTING DATA IN DB
def insert_into_db_data(query):
    engine_url = db_driver + '://' + db_user + ':' + db_user_password + '@' + db_host + ':' + db_port + '/' + db_name
    engine = db.create_engine(engine_url)
    conn = engine.connect()
    conn.execute(query)

# GET FINANCIALS, SEGREGATE THE FINANCIALL DATA AND INSERT IT INTO DB
def write_fin_to_db(findata, table):
    for k, v in findata.items():
        for k1, v1 in v.items():
            if v1 != None:
                for v2 in v1:
                    for k3, v3 in v2.items():
                        df = pd.DataFrame()
                        columns_string = 'insert into ' + table + ' (symbol, date'
                        values_string = 'values ('
                        values_string = values_string + "'" + str(k1) + "'"
                        values_string = values_string + ", '" + str(k3) + "'"
                        for k4, v4 in v3.items():
                            columns_string = columns_string + ', ' + str(k4)
                            if v4 == None:
                                v4 = 'NULL'
                            values_string = values_string + ', ' + str(v4)
                            df[k4] = [v4]
                    columns_string = columns_string + ')'
                    values_string = values_string + ')'
                    insert_query = columns_string + ' ' + values_string
                    print(insert_query)
                    insert_into_db_data(insert_query)
