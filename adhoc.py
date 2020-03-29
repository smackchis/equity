'''
from yahoofinancials import YahooFinancials

yahoo_financials = YahooFinancials('TCS.BO')
ann_report = yahoo_financials.get_financial_stmts('annual', 'income')
qtr_report = yahoo_financials.get_financial_stmts('quaterly', 'income')
'''

import sqlalchemy as db
import pandas as pd

#db_host = '127.0.0.1'
#db_port = '1433'
#db_user = 'wealth_manager'
#db_user_password = 'Rome2017'
#db_table = 'wm'

#engine_url = 'mssql+pymssql://' + db_user + ':' + db_user_password + '@' + db_host + ':' + db_port + '/' + db_table
#print(engine_url)
#engine = db.create_engine(engine_url)
#print(engine.table_names())
#with engine.connect() as conn, conn.begin():
#data = pd.read_sql_table('symbol_list', engine)
#data = pd.read_sql_query('use wm; select symbol from dbo.symbol_list;', engine)
#print(data)

from openpyxl import *

configs_wb = load_workbook(filename="C:\\wm\\equity\\CONFIG\\configs.xlsx",data_only=True)
configs_ws = configs_wb['IND_EQ_DB_Configs']
db_host,db_port,db_table,db_user,db_user_password,db_driver = '','','','','',''
row = 1
while configs_ws.cell(row=row, column= 1).value != None:
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_host':
            db_host = configs_ws.cell(row=row, column=2).value
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_port':
            db_port = configs_ws.cell(row=row, column=2).value
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_user':
            db_user = configs_ws.cell(row=row, column=2).value
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_user_password':
            db_user_password = configs_ws.cell(row=row, column=2).value
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_table':
            db_table = configs_ws.cell(row=row, column=2).value
        if str(configs_ws.cell(row=row, column=1).value).strip() == 'db_driver':
            db_driver = configs_ws.cell(row=row, column=2).value
        row = row+1

print(db_host,db_port,db_table,db_user,db_user_password,db_driver)